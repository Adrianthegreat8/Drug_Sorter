"""
main.py
runs the executable functions
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os
import json
import threading


def save_patient_data(patient_information, file_name='patient_drug_recrod.xlsx'):
    file_path = ensure_path_cwd(file_name)
    workbook = ensure_workbook(file_path)
    add_patient_data(workbook, format_patient_info(patient_information))
    save_workbook(workbook, file_path)
    print(f"Data saved to {file_path}")


# Path functions
def ensure_path_cwd(file_path):
    script_dir = os.path.dirname(os.path.realpath(__file__))
    real_file_path = os.path.join(script_dir, file_path)
    return real_file_path


def ensure_workbook(file_path):
    # Creates a new workbook if the file doesn't exist, else it loads the existing one.
    if os.path.exists(file_path):
        return openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet created with new workbook
        return wb


# Workbook functions
def add_patient_data(workbook, patient_info):
    for drug, dosage_wasted in patient_info['drugs_used'].items():
        if drug not in workbook.sheetnames:  # creates a new worksheet if it does not exist already
            sheet = workbook.create_sheet(title=drug)
            # Add headers as soon as new sheet is created.
            sheet['A1'] = 'Patient Name'
            sheet['B1'] = 'Patient Identifier'
            sheet['C1'] = 'Dosage'
            sheet['D1'] = 'Wasted'
            sheet['E1'] = 'Doctor'
            sheet['F1'] = 'Date'

            next_row = 2  # Start adding data from the second row.
        else:  # selects the correct sheet
            sheet = workbook[drug]
            # Find next empty row (assuming column 1 holds data).
            next_row = sheet.max_row + 1

        alphabet = 'ghijklmnopqrtsuvwxyz'
        assistant_position = []
        for i in range(len(patient_info['assistant(s)'])):
            letter = alphabet[i]
            sheet[letter + '1'] = 'Assistant ' + str(i + 1)
            assistant_position.append(letter + '1')

        # Add patient data.
        sheet.cell(row=next_row, column=1, value=patient_info['patient_name_one_word'])
        sheet.cell(row=next_row, column=2, value=patient_info['patient_identifier'])
        sheet.cell(row=next_row, column=3, value=dosage_wasted[0])
        sheet.cell(row=next_row, column=4, value=dosage_wasted[1])
        sheet.cell(row=next_row, column=5, value=patient_info['doctor'])
        sheet.cell(row=next_row, column=6, value=patient_info['date'])
        i = 7
        for assistant in patient_info['assistant(s)']:
            sheet.cell(row=next_row, column=i, value=assistant)
            i += 1

    auto_adjust_column_width(workbook)


def auto_adjust_column_width(workbook):
    # Auto-adjust column width for readability.
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:  # Needed to avoid error on empty cells.
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width


def save_workbook(workbook, file_path):
    workbook.save(file_path)


def format_patient_info(working_patient_info):
    working_patient_info['patient_name']['First_Name'] = capitalize_first_letter(
        working_patient_info['patient_name']['First_Name'])
    working_patient_info['patient_name']['Last_Name'] = capitalize_first_letter(
        working_patient_info['patient_name']['Last_Name'])

    working_patient_info['patient_name_one_word'] = working_patient_info['patient_name']['First_Name'] + ' ' + \
                                                    working_patient_info['patient_name']['Last_Name']

    new_drugs_used = {}
    for drug_name, amount in working_patient_info['drugs_used'].items():
        formatted_drug_name = capitalize_first_letter(drug_name)  # Capitalize the first letter, rest lowercase
        new_drugs_used[formatted_drug_name] = amount
    working_patient_info['drugs_used'] = new_drugs_used

    formatted_patient_info = working_patient_info

    return formatted_patient_info


def capitalize_first_letter(name):
    first_letter = name[0]
    last_letters = name[1:]
    formatted_name = first_letter.upper() + last_letters
    return formatted_name


class PatientApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.create_window()

        # Read from config.json and set instance variables
        self.config_path = os.path.join(os.path.dirname(__file__), 'config', 'config.json')
        with open(self.config_path, 'r') as config_file:
            config_data = json.load(config_file)
        self.file_name = config_data.get('file_name')
        self.saved_directory = config_data.get('save_directory')

        self.rows = 0
        self.static_stringvars = []  # for storing fixed data like name, identifier
        self.drugs_used = []
        self.assistants = []
        self.create_widgets()
        # self.bind('<Return>', self.handle_return_key)

        self.patient_information = {
            'date': '',
            'patient_name': {'First_Name': '', 'Last_Name': ''},
            'patient_name_one_word': '',
            'patient_identifier': '',
            'drugs_used': {},
            'doctor': '',
            'assistant(s)': []
        }

    def create_window(self):
        # Set the title
        self.title('Patient information system')

        # Set the window size
        window_width, window_height = 600, 800

        # Get the screen dimension
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Find the center position
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)

        # Set the position of the window to the center of the screen
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

    def create_widgets(self):
        self.config_menu()

        # Frame for static fields
        self.static_frame = ttk.Frame(self)
        self.static_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")
        self.static_frame.columnconfigure(1, weight=1)

        # Frame for dynamic fields
        self.dynamic_frame = ttk.Frame(self)
        self.dynamic_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.dynamic_frame.columnconfigure(1, weight=1)

        # Frame for buttons
        self.button_frame = ttk.Frame(self)
        self.button_frame.grid(row=2, column=0, padx=10, pady=5)

        self.grid_columnconfigure(0, weight=1)  # Ensures the main window expands content

        # Initialize static fields
        self.init_static_fields()

        # Initialize buttons
        self.init_buttons()

    def config_menu(self):
        menu_bar = tk.Menu(self)
        self.config(menu=menu_bar)
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Set Save Directory", command=self.set_saved_directory)
        file_menu.add_command(label="Submit", command=self.submit_button_command)

    def set_saved_directory(self):
        # Open a dialog to select a directory
        directory = filedialog.askdirectory()
        if directory:
            self.saved_directory = directory
            self.update_config_attribute('save_directory', directory)
            messagebox.showinfo("Directory Set", f"Files will be saved in: {self.saved_directory}")
            # write to config.json new default directory

    def load_config(self):
        if not os.path.exists(self.config_path):
            raise FileNotFoundError(f"Config file not found at {self.config_path}")
        with open(self.config_path, 'r') as config_file:
            return json.load(config_file)

    def save_config(self, config_data):
        with open(self.config_path, 'w') as config_file:
            json.dump(config_data, config_file, indent=2)

    def update_config_attribute(self, key, value):
        # Load the existing configuration
        config = self.load_config()

        # Update the desired attribute
        config[key] = value

        # Save the updated configuration back to config.json
        self.save_config(config)
        print(f"Updated {key} to {value}")

    def init_static_fields(self):
        self.first_name_var = tk.StringVar()
        self.first_name_entry = self.create_entry_field("First Name:", 0, self.first_name_var, self.static_frame)

        self.last_name_var = tk.StringVar()
        self.last_name_entry = self.create_entry_field("Last Name:", 1, self.last_name_var, self.static_frame)

        self.identifier_var = tk.StringVar()
        self.identifier_entry = self.create_entry_field("Identifier:", 2, self.identifier_var, self.static_frame)

        self.date_var = tk.StringVar()
        self.date_entry = self.create_entry_field('Date:', 3, self.date_var, self.static_frame)

        self.doctor_var = tk.StringVar()
        self.doctor_entry = self.create_entry_field('Doctor:', 4, self.doctor_var, self.static_frame)

    def create_entry_field(self, label_text, row, variable, frame):
        ttk.Label(frame, text=label_text).grid(row=row, column=0, padx=10, pady=5, sticky="ew")
        self.rows += 1
        entry = ttk.Entry(frame, textvariable=variable)
        entry.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
        self.static_stringvars.append(variable)
        return entry, variable  # Return both the entry widget and its associated StringVar

    def init_buttons(self):
        self.add_drug_button = ttk.Button(self.button_frame, text="Add Drug", command=self.add_drug_field)
        self.add_drug_button.grid(row=0, column=0, pady=10, padx=10, sticky='ew')

        self.add_assistant_button = ttk.Button(self.button_frame, text="Add Assistant",
                                               command=self.add_assistant_field)
        self.add_assistant_button.grid(row=0, column=1, pady=10, padx=10, sticky='ew')

        self.submit_button = ttk.Button(self.button_frame, text="Submit", command=self.submit_button_command)
        self.submit_button.grid(row=0, column=2, pady=10, padx=10, sticky="ew")

    def add_drug_field(self):
        # Increment the row for new fields
        current_row = self.rows
        self.rows += 2  # Prepare for the next widget (we're adding two fields per drug)

        # Create variables to store drug name and dosage input
        drug_var = tk.StringVar()
        dosage_var = tk.StringVar()
        wasted_var = tk.StringVar()

        # Create and place the drug name label and entry field
        ttk.Label(self.dynamic_frame, text=f"Drug #{len(self.drugs_used) + 1}:").grid(row=current_row, column=0,
                                                                                      padx=10, pady=5, sticky="we")
        drug_entry = ttk.Entry(self.dynamic_frame, textvariable=drug_var)
        drug_entry.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")

        # Create and place the dosage label and entry field
        ttk.Label(self.dynamic_frame, text="Dosage:").grid(row=current_row + 1, column=0, padx=10, pady=5, sticky="we")
        dosage_entry = ttk.Entry(self.dynamic_frame, textvariable=dosage_var)
        dosage_entry.grid(row=current_row + 1, column=1, padx=10, pady=5, sticky="ew")

        # Create and place the wasted label and entry field
        ttk.Label(self.dynamic_frame, text="Wasted:").grid(row=current_row + 2, column=0, padx=10, pady=5, sticky="we")
        wasted_entry = ttk.Entry(self.dynamic_frame, textvariable=wasted_var)
        wasted_entry.grid(row=current_row + 2, column=1, padx=10, pady=5, sticky="ew")

        # Store the drug name and dosage entries for later use (if needed)
        self.drugs_used.append((drug_var, dosage_var, wasted_var))
        self.rows += 1  # Increment rows to make space for the "Add Drug" button

    def add_assistant_field(self):
        # Increment the row for new fields
        current_row = self.rows
        self.rows += 1  # Prepare for the next widget (we're adding two fields per drug)

        # Create variables to store drug name and dosage input
        assistant = tk.StringVar()

        # Create and place the drug name label and entry field
        ttk.Label(self.dynamic_frame, text=f"Assistant #{len(self.assistants) + 1}:").grid(row=current_row, column=0,
                                                                                           padx=10,
                                                                                           pady=5, sticky="we")
        assistant_entry = ttk.Entry(self.dynamic_frame, textvariable=assistant)
        assistant_entry.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")

        # Store the assistant name  for later use (if needed)
        self.assistants.append(assistant)

    def submit_button_command(self):
        if not self.saved_directory:  # This is placed here so that fields are not cleared if a dir is not set
            messagebox.showerror("Error", "Please set a save directory first.")
            return
        self.submit_patient_info()
        self.clear_fields()
        self.reset_fields()

    def reset_fields(self):
        self.reset_drug_field()
        self.reset_assistant_field()
        self.reset_dynamic_frame()
        self.update_idletasks()

    def reset_drug_field(self):
        self.rows = 0
        self.drugs_used.clear()

    def reset_assistant_field(self):
        self.rows = 0
        self.assistants.clear()

    def clear_fields(self):
        # clear static fields
        for i in self.static_stringvars:
            i.set('')

        for i in self.assistants:
            i.set('')

        for i in self.drugs_used:
            for j in i:
                j.set('')

    def reset_dynamic_frame(self):
        for widget in self.dynamic_frame.winfo_children():
            widget.destroy()
        self.dynamic_frame.destroy()
        self.dynamic_frame = ttk.Frame(self)
        self.dynamic_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.dynamic_frame.columnconfigure(1, weight=1)

    def submit_patient_info(self):
        static_variable_list = [i.get() for i in self.static_stringvars]
        assistants_list = [i.get() for i in self.assistants]
        drugs_used_dict = {drug.get(): [dosage.get(), wasted.get()] for drug, dosage, wasted in self.drugs_used}

        self.patient_information = {
            'date': static_variable_list[3],
            'patient_name': {'First_Name': static_variable_list[0], 'Last_Name': static_variable_list[1]},
            'patient_name_one_word': '',
            'patient_identifier': static_variable_list[2],
            'drugs_used': drugs_used_dict,
            'doctor': static_variable_list[4],
            'assistant(s)': assistants_list
        }

        full_file_path = os.path.join(self.saved_directory, self.file_name + '.xlsx')
        # debugging
        # print('saved dir:', self.saved_directory)
        # print('file name:', self.file_name)
        # print('Saving to:', full_file_path)
        # print(type(full_file_path))
        threading.Thread(target=save_patient_data, args=(self.patient_information, full_file_path)).start()

    def periodic_update(self):
        self.after(50, self.periodic_update)


app = PatientApp()
app.mainloop()
