import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os


# Path functions
def ensure_path_cwd(file_path):
    script_dir = os.path.dirname(os.path.realpath(__file__))
    real_file_path = os.path.join(script_dir, file_path)
    return real_file_path

def ensure_workbook(file_path):
    # Creates a new workbook if the file doesn't exist, else it loads the existing one.
    if os.path.exists(file_path):
        return openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet created with new workbook
        return wb

# Worbook functions
def add_patient_data(workbook, patient_info):
    for drug, dosage in patient_info['drugs_used'].items():
        if drug not in workbook.sheetnames: #creates a new worksheet if it does not exist already
            sheet = workbook.create_sheet(title=drug)
            # Add headers as soon as new sheet is created.
            sheet['A1'] = 'Patient Name'
            sheet['B1'] = 'Patient Identifier'
            sheet['C1'] = 'Dosage'
            next_row = 2  # Start adding data from the second row.
        else: #selects the correct sheet
            sheet = workbook[drug]
            # Find next empty row (assuming column 1 holds data).
            next_row = sheet.max_row + 1

        # Add patient data.
        sheet.cell(row=next_row, column=1, value=patient_info['patient_name_one_word'])
        sheet.cell(row=next_row, column=2, value=patient_info['patient_identifier'])
        sheet.cell(row=next_row, column=3, value=dosage)

    auto_adjust_column_width(workbook)

def auto_adjust_column_width(workbook):
    # Auto-adjust column width for readability.
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:  # Needed to avoid error on empty cells.
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def save_workbook(workbook, file_path):
    workbook.save(file_path)

# Data functions
def enter_patient_info():
    working_patient_info = {
        'date': '',
        'patient_name': {'First_Name': '', 'Last_Name': ''},
        'patient_name_one_word': '',
        'patient_identifier': '',
        'drugs_used': {},
        'doctor': '',
        'assistant(s)': []
    }
    for i in working_patient_info:
        if i == 'patient_name':
            print('Enter', i + '\n')
            patient_name = input()

            if ' ' in patient_name:
                patient_name_split = patient_name.split()
                working_patient_info['patient_name']['First_Name'] = patient_name_split[0]
                working_patient_info['patient_name']['Last_Name'] = patient_name_split[1]
            else:
                working_patient_info['patient_name']['First_Name'] = patient_name
        elif i == 'assistant(s)':
            num_assistaints = int(input('How many assistants? \n'))
            for i in range(num_assistaints):
                print('Enter assistant', i+1)
                assistant_name = input()
                working_patient_info['assistant(s)'].append(assistant_name)
        elif i == 'drugs_used':
            num_drugs_used = int(input('How many drugs used? \n'))
            for i in range(num_drugs_used):
                print('Enter drug', str(i+1)+ '\n')
                drug_name = input()
                ammount_used = input('Enter ammount administered \n')
                working_patient_info['drugs_used'][drug_name] = ammount_used
        elif i == 'patient_name_one_word':
            continue
        else:
            print('Enter', i + '\n')
            patient_item = input()
            working_patient_info[i] = patient_item

    formatted_patient_info = format_patient_info(working_patient_info)

    formatted_patient_info['patient_name_one_word'] = formatted_patient_info['patient_name']['First_Name']+' '+formatted_patient_info['patient_name']['Last_Name']

    return formatted_patient_info

def format_patient_info(working_patient_info):

    working_patient_info['patient_name']['First_Name'] = capitalize_first_letter(working_patient_info['patient_name']['First_Name'])
    working_patient_info['patient_name']['Last_Name'] = capitalize_first_letter(working_patient_info['patient_name']['Last_Name'])

    new_drugs_used = {}
    for drug_name, amount in working_patient_info['drugs_used'].items():
        formatted_drug_name = capitalize_first_letter(drug_name)  # Capitalize the first letter, rest lowercase
        new_drugs_used[formatted_drug_name] = amount
    working_patient_info['drugs_used'] = new_drugs_used

    formatted_patient_info = working_patient_info

    return formatted_patient_info

def capitalize_first_letter(name):
    first_letter = name[0]
    last_letters = name[1:]
    formatted_name = first_letter.upper() + last_letters
    return formatted_name

# Main 
def main():
    # Sample patient data.
    sample_patient_info = {
        'patient_name': {'First_Name': 'michael', 'Last_Name': 'prieto'},
        'patient_name_one_word': '',
        'patient_identifier': '123456',
        'drugs_used': {
            'aspirin': '100mg',
            'ibuprofen': '200mg'
        },
        'doctor': 'Ibanez',
        'assistant(s)' : ['Alex']
    }

    file_path = 'patient_drug_record.xlsx'
    file_path = ensure_path_cwd(file_path)

    workbook = ensure_workbook(file_path)

    add_patient_data(workbook, enter_patient_info())

    save_workbook(workbook, file_path)
    print(f"Data saved to {file_path}")

if __name__ == '__main__':
    main()