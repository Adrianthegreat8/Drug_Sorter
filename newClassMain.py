import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os

class Patient:
    def __init__(self, date, first_name, last_name, identifier, doctor, assistants=[]):
        self.date = date
        self.first_name = first_name.capitalize()
        self.last_name = last_name.capitalize()
        self.identifier = identifier
        self.drugs_used = {}  # Dictionary to store drug information
        self.doctor = doctor
        self.assistants = assistants

    def add_drug(self, drug_name, amount_administered):
        self.drugs_used[drug_name.capitalize()] = amount_administered

def ensure_workbook(file_path):
    if os.path.exists(file_path):
        return openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        return wb


def type_to_string(input): #converts type function into a string output
    output = ''.join(list(str(type(input)))[8:-2])
    return output

def ensure_datatype(data_input, type): #checks the type of an input, returns bool
    if type_to_string(data_input) == type:
        return True
    else:
        return False


def add_patient_data(workbook, patient):
    for drug, dosage in patient.drugs_used.items():
        if drug not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=drug)
            sheet['A1'] = 'Date'
            sheet['B1'] = 'Patient Name'
            sheet['C1'] = 'Patient Identifier'
            sheet['D1'] = 'Dosage'
            sheet['E1'] = 'Doctor'
            sheet['F1'] = 'Assistants'
            next_row = 2
        else:
            sheet = workbook[drug]
            next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1, value=patient.date)
        sheet.cell(row=next_row, column=2, value=f"{patient.first_name} {patient.last_name}")
        sheet.cell(row=next_row, column=3, value=patient.identifier)
        sheet.cell(row=next_row, column=4, value=dosage)
        sheet.cell(row=next_row, column=5, value=patient.doctor)
        sheet.cell(row=next_row, column=6, value=', '.join(patient.assistants))

def auto_adjust_column_width(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def save_workbook(workbook, file_path):
    workbook.save(file_path)

def enter_patient_info():
    print("Please enter patient information:")
    date = input("Date (DD-MM-YYYY): ")

    while True:
        first_name = input("First name: ")
        if ensure_datatype(first_name, 'str') == True:
            break
        else:
            print('First name must be a string. Try again\n')

    while True:
        last_name = input("Last name: ")
        if ensure_datatype(last_name, 'str') == True:
            break
        else:
            print('Last name must be a string. Try again\n')

    while True:
        identifier = input("Patient identifier: ")
        if ensure_datatype(identifier, 'int') == True:
            break
        else:
            print('Patient identifier must be a number. Try again\n')

    while True:
        doctor = input("Doctor's name: ")
        if ensure_datatype(doctor, 'str') == True:
            break
        else:
            print("Doctor's name must be a string of characters. Try again\n")

    assistants = []

    while True:
        num_assistants = int(input("How many assistants? "))
        if ensure_datatype(num_assistants, 'int') == True:
            break
        else:
            print('Number of assistants must be a number. Try again\n')

    for i in range(num_assistants):
        assistant_name = input(f"Enter assistant {i+1} name: ")
        assistants.append(assistant_name)

    patient = Patient(date, first_name, last_name, identifier, doctor, assistants)

    while True:
        num_drugs_used = int(input("How many drugs used? "))
        if ensure_datatype(num_drugs_used, 'int') == True:
            break
        else:
            print('Number of drugs used must be a number. Try again\n')

    for i in range(num_drugs_used):
        drug_name = input(f"Enter drug {i+1} name: ")
        amount_administered = input("Amount administered: ")
        patient.add_drug(drug_name, amount_administered)

    return patient

def main():
    file_path = 'patient_drug_record.xlsx'
    workbook = ensure_workbook(file_path)

    patient = enter_patient_info()
    add_patient_data(workbook, patient)

    auto_adjust_column_width(workbook)
    save_workbook(workbook, file_path)
    print(f"Data saved to {file_path}")

if __name__ == '__main__':
    main()
